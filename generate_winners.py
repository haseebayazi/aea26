#!/usr/bin/env python3
"""
Generate AEA26 Winners Excel file and organized Winner folders.
"""

import os
import glob
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = "/home/user/aea26"

# ── Winners definition ──────────────────────────────────────────────────────
WINNERS = {
    "Professional Achievement": [
        {"rank": 1, "id": 10,  "name": "Anum Ali"},
        {"rank": 2, "id": 190, "name": "Syed Muzahir Abbas"},
        {"rank": 3, "id": 15,  "name": "Muhammad Usman Sheikh"},
    ],
    "Distinguished Young Alumni": [
        {"rank": 1, "id": 60,  "name": "Maryam Arif"},
        {"rank": 2, "id": 180, "name": "Dr Rehmat Ullah"},
        {"rank": 3, "id": 109, "name": "SAKINA FATIMA"},
    ],
    "Innovation & Entrepreneurship": [
        {"rank": 1, "id": 33,  "name": "Farah Gul Rahuja"},
        {"rank": 2, "id": 27,  "name": "Muhammad Salman Latif"},
        {"rank": 3, "id": 132, "name": "Junaid"},
    ],
    "Social Impact & Community Service": [
        {"rank": 1, "id": 177, "name": "Muhammad Mursaleen"},
        {"rank": 2, "id": 49,  "name": "Muhammad Asim Bhatti"},
        {"rank": 3, "id": 130, "name": "Malik Waleed Paracha"},
    ],
}

# Source candidate folders (directory names contain literal &amp; not &)
SOURCE_DIRS = {
    "Professional Achievement":          os.path.join(BASE_DIR, "1-Professional Achievement"),
    "Distinguished Young Alumni":        os.path.join(BASE_DIR, "2-Distinguished Young Alumni"),
    "Innovation & Entrepreneurship":     os.path.join(BASE_DIR, "3-Innovation &amp; Entrepreneurship"),
    "Social Impact & Community Service": os.path.join(BASE_DIR, "4-Social Impact &amp; Community Service"),
}

# Header colours (ARGB hex, no leading #)
CATEGORY_COLORS = {
    "Professional Achievement":          "FF4472C4",  # Blue
    "Distinguished Young Alumni":        "FF7030A0",  # Purple
    "Innovation & Entrepreneurship":     "FF538135",  # Green
    "Social Impact & Community Service": "FFC00000",  # Red
}

RANK_LABELS = {1: "1st", 2: "2nd", 3: "3rd"}

# Columns to include in output (order matters)
OUTPUT_COLUMNS = [
    "Rank",
    "Name", "Father", "RegNo", "Department", "Campus",
    "CNIC", "DegreeProgram", "ProfessionalTitle", "Organization",
    "Email", "Phone", "LinkedInUrl",
    "Career Score", "Career Brief",
    "Outcomes Score", "Outcomes Brief",
    "Societal Score", "Societal Brief",
    "Projects Score", "Projects Brief",
    "Leadership Score", "Leadership Brief",
    "Mentoring Score", "Mentoring Brief",
    "Service Score", "Service Brief",
    "Inspiration Score", "Inspiration Brief",
    "Entrepreneur Score", "Entrepreneur Brief",
    "Research Score", "Research Brief",
    "Patent Score", "Patent Brief",
    "Creativity Score", "Creativity Brief",
    "Integrity Score", "Integrity Brief",
    "Alumni Score", "Alumni Brief",
    "Community Score", "Community Brief",
    "Reputation Score", "Reputation Brief",
]


# ── Load xlsx data ───────────────────────────────────────────────────────────

def find_xlsx(pattern):
    matches = glob.glob(os.path.join(BASE_DIR, pattern))
    if not matches:
        raise FileNotFoundError(f"No file matching {pattern}")
    return matches[0]


def load_sheet_data(filepath):
    """
    Load the sheet with the most score-column data.
    Returns (headers_list, {sr_no: row_dict}).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    best_sheet = None
    best_score_cols = -1

    for sname in wb.sheetnames:
        ws = wb[sname]
        headers = [cell.value for cell in ws[1]]
        score_cols = sum(1 for h in headers if h and "Score" in str(h))
        print(f"  Sheet '{sname}': {len(headers)} cols, {score_cols} score cols")
        if score_cols > best_score_cols:
            best_score_cols = score_cols
            best_sheet = sname

    print(f"  → Using sheet: '{best_sheet}'")
    ws = wb[best_sheet]
    headers = [cell.value for cell in ws[1]]

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sr_no = row[0]
        if sr_no is None:
            continue
        try:
            sr_no = int(sr_no)
        except (ValueError, TypeError):
            continue
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if headers[i] is not None}
        data[sr_no] = row_dict

    return headers, data


def load_all_category_data():
    files = {
        "Professional Achievement":          find_xlsx("Professional Achievement*.xlsx"),
        "Distinguished Young Alumni":        find_xlsx("Distinguished Young Alumni*.xlsx"),
        "Innovation & Entrepreneurship":     find_xlsx("Innovation*.xlsx"),
        "Social Impact & Community Service": find_xlsx("Social Impact*.xlsx"),
    }

    all_data = {}
    for cat, fpath in files.items():
        print(f"\nLoading: {fpath}")
        headers, data = load_sheet_data(fpath)
        all_data[cat] = data

    return all_data


# ── Excel creation ───────────────────────────────────────────────────────────

def create_excel(all_data):
    out_path = os.path.join(BASE_DIR, "AEA26 Winners - Candidate Details.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for cat, winners in WINNERS.items():
        ws = wb.create_sheet(title=cat)
        color_hex = CATEGORY_COLORS[cat]
        fill = PatternFill(fill_type="solid", fgColor=color_hex)
        header_font = Font(bold=True, color="FFFFFFFF", size=11)
        thin_side = Side(style="thin", color="FFCCCCCC")
        thin_border = Border(
            left=thin_side, right=thin_side,
            top=thin_side, bottom=thin_side,
        )

        # Write header row
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border = thin_border

        cat_data = all_data.get(cat, {})

        # Write winner rows
        for row_idx, winner in enumerate(winners, start=2):
            cand_id = winner["id"]
            row_data = cat_data.get(cand_id, {})
            if not row_data:
                print(f"  WARNING: No data found for {cat} ID {cand_id} ({winner['name']})")

            for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
                if col_name == "Rank":
                    val = RANK_LABELS[winner["rank"]]
                else:
                    val = row_data.get(col_name, "")

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Alternate row shading
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFF2F2F2")

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-width columns (cap at 60)
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            for row_idx in range(2, len(winners) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    for line in str(val).split("\n"):
                        max_len = max(max_len, len(line))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        # Row height for data rows
        ws.row_dimensions[1].height = 20
        for row_idx in range(2, len(winners) + 2):
            ws.row_dimensions[row_idx].height = 45

        print(f"  Sheet '{cat}' created with {len(winners)} winner rows.")

    wb.save(out_path)
    print(f"\n✓ Excel file saved: {out_path}")
    return out_path


# ── Winner folders ───────────────────────────────────────────────────────────

def find_source_folder(source_base, cand_id):
    """Find a candidate folder by numeric ID prefix."""
    prefix = f"{cand_id}-"
    try:
        entries = os.listdir(source_base)
    except FileNotFoundError:
        return None
    for entry in entries:
        if entry.startswith(prefix) and os.path.isdir(os.path.join(source_base, entry)):
            return os.path.join(source_base, entry)
    return None


def create_winner_folders():
    winners_base = os.path.join(BASE_DIR, "Winners")
    os.makedirs(winners_base, exist_ok=True)
    print(f"\nCreating winner folders under: {winners_base}")

    total_files = 0
    total_folders = 0

    for cat, winners in WINNERS.items():
        cat_dir = os.path.join(winners_base, cat)
        os.makedirs(cat_dir, exist_ok=True)
        print(f"\n  Category: {cat}")

        source_base = SOURCE_DIRS[cat]

        for winner in winners:
            rank_label = RANK_LABELS[winner["rank"]]
            dest_folder_name = f"{rank_label} - {winner['name']}"
            dest_dir = os.path.join(cat_dir, dest_folder_name)
            os.makedirs(dest_dir, exist_ok=True)
            total_folders += 1

            src_dir = find_source_folder(source_base, winner["id"])
            if not src_dir:
                print(f"    WARNING: Source folder not found for ID {winner['id']} ({winner['name']})")
                continue

            # Copy all files
            files_copied = 0
            for fname in os.listdir(src_dir):
                src_file = os.path.join(src_dir, fname)
                if os.path.isfile(src_file):
                    shutil.copy2(src_file, os.path.join(dest_dir, fname))
                    files_copied += 1
                    total_files += 1

            print(f"    {rank_label} - {winner['name']} (ID {winner['id']}): "
                  f"copied {files_copied} file(s) from {os.path.basename(src_dir)}")

    print(f"\n✓ Winner folders created: {total_folders} folders, {total_files} files copied.")
    return winners_base


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("AEA26 Winners Generator")
    print("=" * 60)

    print("\n[1/3] Loading candidate data from xlsx files...")
    all_data = load_all_category_data()

    print("\n[2/3] Creating Excel file...")
    excel_path = create_excel(all_data)

    print("\n[3/3] Creating winner folders...")
    winners_base = create_winner_folders()

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Excel file : {excel_path}")
    print(f"Winner dirs: {winners_base}/")
    print()
    for cat, winners in WINNERS.items():
        print(f"  {cat}:")
        for w in winners:
            print(f"    {RANK_LABELS[w['rank']]} place: {w['name']} (ID #{w['id']})")
    print()
    print("Done!")


if __name__ == "__main__":
    main()
